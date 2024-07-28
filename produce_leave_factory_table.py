import fitz  # PyMuPDF
from openpyxl import load_workbook

# 用于返回所需要的数据
def obtain_data(input_pdf_path):
    # 所有数据
    total_data = dict()
    doc = fitz.open(input_pdf_path)
    for page in doc:
        text_instances = page.get_text("dict")["blocks"]  # 获取页面上的所有文字字典
        for inst in text_instances:
            # 判断是不是表格、水印
            if inst.get('lines') and inst['lines'][0]['spans'][0]['color'] != 4144959:
                # 每个城市对应一行，这里存储的是城市那一列的x轴位置
                city_cell = inst['lines'][0]['spans'][0]
                # 如果是城市的位置那就代表一整行都是需要的数据
                if city_cell['origin'][0] == 29.670000076293945:
                    # 这里存的是城市名称
                    city_name = city_cell['text']
                    # 定义字典来存储各个城市的价格数据
                    total_data[city_name] = dict()
                    for i in inst['lines'][1:]:
                        # 使用get来判断数据是否在字典
                        temp = data_dict.get(i['spans'][0]['origin'][0], False)
                        if temp:
                            total_data[city_name][temp] = i['spans'][0]['text']
    return total_data

data_dict = {
    29.670000076293945: '城市',
    218.25: 'HRB400E螺纹钢Ф18X12',
    264.0799865722656: 'HRB400E盘螺Ф8.0',
    309.9200134277344: 'HPB300线材Ф8'
}

# 设置路径
filename = r"C:\Users\Administrator\Desktop\路径.txt"

with open(filename, 'r', encoding='utf-8') as file:
    lines = file.readlines()

pdf_path, template_path, out_path = [i.strip().split("=")[1] for i in lines]


# 获取pdf的数据
pdf_data = obtain_data(pdf_path)
wb = load_workbook(template_path)
ws = wb.active

# 这三个位置向右边3格就是要填充的数据
location = {"Ф8.0": "HRB400E盘螺Ф8.0", "Ф18X12": "HRB400E螺纹钢Ф18X12", "Ф8": "HPB300线材Ф8"}

for col in ws.iter_cols(max_row=20, min_col=1):
    for row in col:
        if type(row.value) is str and row.value.count("粤裕丰出厂价表（2024版）") >= 1 and pdf_data.get(row.value[:3]):
            temp = pdf_data.get(row.value[:3])
            row.offset(row=4, column=4).value = temp['HRB400E盘螺Ф8.0']  # HRB400E盘螺
            row.offset(row=11, column=4).value = temp['HRB400E螺纹钢Ф18X12']  # HRB400E螺纹钢
            row.offset(row=18, column=4).value = temp['HPB300线材Ф8']  # HPB300线材

wb.save(out_path)
